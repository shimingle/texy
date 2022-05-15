# !/usr/bin/env python
# -*- coding: UTF-8 -*-

"""
@Project ：Test
@File    ：疫情信息获取.py
@Version ：v1.0
@Author  ：李文竹
@Date    ：2022/5/13 18:37
@Desc    : 
"""

import os
import shutil
import getpass
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait


class Excel(object):
    """
    Excel文件处理
    """

    def __init__(self):
        self.filename = None
        self.workbook = None
        self.worksheet = None

    def load(self, filename, sheet_name=None):
        """
        读取excel文件
        :param filename: Excel文件名
        :param sheet_name: sheet名
        """
        self.filename = filename

        if os.path.exists(filename):
            self.workbook = openpyxl.load_workbook(filename)
        else:
            self.workbook = openpyxl.Workbook()

        if sheet_name is None:
            sheet_name = self.workbook.sheetnames[0]

        self.worksheet = self.workbook[sheet_name]

    def save(self, filename=None):
        """
        保存Excel
        :param filename: 文件名
        """
        if filename is None:
            filename = self.filename

        self.workbook.save(filename)

    def close(self):
        """
        关闭Excel
        :param filename: 文件名
        """
        self.workbook.close()


class Browser(object):
    """
    浏览器工具类
    """

    def __init__(self):
        self.driver = None

    def web_driver(self):
        """
        浏览器驱动
        :return: driver
        """
        options = webdriver.ChromeOptions()
        options.add_experimental_option('excludeSwitches', ['enable-automation'])
        options.add_experimental_option('useAutomationExtension', False)
        driver = webdriver.Chrome(options=options)
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                Object.defineProperty(navigator, 'webdriver', {
                  get: () => undefined
                })
              """
        })
        driver.implicitly_wait(30)
        driver.maximize_window()
        self.driver = driver

    def presence_of_element_located(self, xpath, times=120):
        """
        等待元素被添加到dom里
        :param xpath: xpath
        :param times: 超时时间
        :return: 元素对象
        """
        try:
            element = WebDriverWait(self.driver, times).until(
                expected_conditions.presence_of_element_located((By.XPATH, xpath)))
            return element
        except Exception as e:
            print("等待元素被添加到dom里失败：%s" % e)
            return

    def visibility_of_element_located(self, xpath, times=120):
        """
        等待元素被添加到dom里并且可见
        :param xpath: xpath
        :param times: 超时时间
        :return: 元素对象
        """
        try:
            element = WebDriverWait(self.driver, times).until(
                expected_conditions.visibility_of_element_located((By.XPATH, xpath)))
            return element
        except Exception as e:
            print("等待元素被添加到dom里并且可见失败：%s" % e)
            return

    def element_to_be_clickable(self, xpath, times=300):
        """
        等待元素可见并且是enable状态，代表可点击
        :param xpath: xpath
        :param times: 超时时间
        :return: 元素对象
        """
        try:
            element = WebDriverWait(self.driver, times).until(
                expected_conditions.element_to_be_clickable((By.XPATH, xpath)))
            return element
        except Exception as e:
            print("等待元素可见并且是enable状态失败：%s" % e)
            return


class Process(object):
    """
    业务处理
    """
    def __init__(self):
        self._browser = Browser()
        self._excel = Excel()

    @staticmethod
    def _del_file(file_or_dir, startswith=str(), endswith=str()):
        """
        删除文件
        @param file_or_dir: 文件或目录
        @param startswith: 文件前缀
        @param endswith: 文件后缀
        @return: None
        """
        try:
            if os.path.isfile(file_or_dir):
                os.remove(file_or_dir)
                print('删除文件：' + file_or_dir)
            elif os.path.isdir(file_or_dir):
                for filename in os.listdir(file_or_dir):
                    file = os.path.join(file_or_dir, filename)
                    if os.path.isfile(file) and filename.startswith(startswith) and filename.endswith(endswith):
                        os.remove(file)
                        print('删除文件：' + file)
                    else:
                        shutil.rmtree(file, ignore_errors=True)
                        print('删除文件夹：' + file)
        except Exception as e:
            print(e)

    def business(self, result_file):
        """
        业务调度
        @param result_file: 结果文件
        @return: None
        """
        try:
            self._del_file(result_file)
            self._excel.load(result_file)
            self._browser.web_driver()
            url = 'https://voice.baidu.com/act/newpneumonia/newpneumonia/?from=osari_aladin_banner#tab4'
            self._browser.driver.get(url)
            xpath = '//*[@id="nationTable"]/table'
            epidemic_summary = self._browser.visibility_of_element_located(xpath)
            tr_ele = epidemic_summary.find_elements(by=By.TAG_NAME, value='tr')
            for tr_index in range(len(tr_ele)):
                temp_list = []
                if tr_index == 0:
                    th_ele = tr_ele[tr_index].find_elements(by=By.TAG_NAME, value='th')
                    for ele in th_ele:
                        text = ele.find_element(by=By.TAG_NAME, value='span').text
                        temp_list.append(text)
                else:
                    td_ele = tr_ele[tr_index].find_elements(by=By.TAG_NAME, value='td')
                    for td_index in range(len(td_ele)):
                        if td_index == 0:
                            text = td_ele[td_index].find_elements(by=By.TAG_NAME, value='span')[-1].text
                        else:
                            text = td_ele[td_index].text
                        temp_list.append(text)

                print(temp_list)
                self._excel.worksheet.append(temp_list)
        except Exception as e:
            print(e)
        finally:
            self._browser.driver.quit()
            self._excel.save()
            self._excel.close()


def main():
    """
    业务发起
    """
    result_file = r"C:\Users\{user}\Desktop\疫情信息.xlsx".format(user=getpass.getuser())
    result_file = os.path.abspath(result_file)
    Process().business(result_file)


main()
